import openpyxl
import openpyxl.worksheet as worksheet

from file_features import output_message_exit, file_unused


class ExcelFileControl:
    """ Класс для управления файлом excel используя пакет openpyxl.  """

    def __init__(self, full_file_name: str = None):
        self.file_name = full_file_name
        self.sheet_name = None
        self.book = None
        self.sheet = None
        self.open_file()

    def __enter__(self):
        """ Вызывается при старте контекстного менеджера with. """
        return self if self.book else None

    def __exit__(self, exception_type, exception_value, traceback):
        """ Будет вызван в завершении конструкции with, или в случае возникновения исключения. """
        self.close_file()

    def __str__(self):
        return f"excel file: '{self.file_name}', sheet: '{self.sheet_name}', link_sheet: '{self.sheet}'"

    def open_file(self):
        """ Создает объект 'excel файл'. """
        try:
            self.book = openpyxl.Workbook()  # создаем экземпляр класса
        except IOError as err:
            output_message_exit(f"Ошибка при создании файла: {self.file_name!r}", f"{err}")

    def close_file(self):
        """ Записывает объект Workbook() в файл"""
        if self.book:
            if file_unused(self.file_name):
                self.book.save(self.file_name)
                self.book.close()
                self.sheet_name = None
                self.book = None
                self.sheet = None
            else:
                output_message_exit(f"ошибка записи в файл: {self.file_name!r}",
                                    f"файл используется другим приложением.")
        else:
            output_message_exit(f"ошибка записи в файл: {self.file_name!r}",
                                f"нет данных для записи.")

    def create_sheets(self, sheets_name: list[str] = None):
        """ Создает листы в книге. Если в книге есть другие листы удаляет их. """
        if self.book and sheets_name:
            for name in sheets_name:
                self.book.create_sheet(name)
            for sheet_i in self.book.worksheets:
                if sheet_i.title not in sheets_name:
                    self.book.remove(sheet_i)

    def get_sheet(self, sheet_name: str) -> worksheet:
        """ Получает ссылку на страницу с именем sheet_name  """
        if sheet_name and self.book and sheet_name in self.book.sheetnames:
            self.sheet_name = sheet_name
            self.sheet = self.book[sheet_name]
            return self.sheet

    def set_grid(self, sheet_name: str = None, grid: bool = True):
        """ Отключает/включает сетку на листе sheet """
        target_sheet = self.get_sheet(sheet_name)
        if target_sheet:
            target_sheet.sheet_view.showGridLines = grid


if __name__ == "__main__":
    full_name = r"F:\Kazak\GoogleDrive\1_KK\Job_CNAC\1_targets\SRC\11-09-2023\output_1\Test_Output.xlsx"
    file = ExcelFileControl(full_name)
    with file as ex:
        sheets = ["report", "stat"]
        ex.create_sheets(sheets)
        wn = sheets[0]
        sheet = ex.get_sheet(wn)
        ex.set_grid(ex.sheet.title, True)
        ex.set_grid(wn, False)
    # ex.close_file()
