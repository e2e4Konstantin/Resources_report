from pandas import DataFrame, isna
import numpy
from openpyxl.worksheet import worksheet
from openpyxl.utils.cell import column_index_from_string

from data_frame_features import SourceData
from layout.layout_setting import headers, items_styles
from layout.write_resources_table import write_resources_for_table
from layout.set_cell_style import set_cell_style
from layout.layout_setting import basic_colors

import time


def _attributes_output(src_attributes: str, start_column: int, sheet: worksheet, row: int) -> int:
    if not isna(src_attributes):
        sheet.cell(row=row, column=start_column).value = headers['attribute'][0]
        for column in range(start_column, start_column + len(headers["attribute"])):
            set_cell_style(sheet, row, column, 'attributes_header')

        attributes = src_attributes.split(',')
        attributes_length = len(attributes)
        for i, attribute in enumerate(attributes):
            column = start_column + i
            sheet.cell(row=row + 1, column=column).value = attribute
            set_cell_style(sheet, row + 1, column, 'attributes_volume')
            set_cell_style(sheet, row, column, 'attributes_header')

        if attributes_length > 1:
            sheet.merge_cells(start_row=row, start_column=start_column,
                              end_row=row, end_column=start_column + attributes_length - 1)
        return start_column + attributes_length
    return start_column


def _parameters_output(src_parameters: str, start_column: int, sheet: worksheet, row: int):
    if not isna(src_parameters):
        parameters = src_parameters.split(',')
        # print(parameters)
        parameters_length = len(parameters)
        if parameters_length > 0:
            mini_table_header = headers['option']  # таблица: 'от', 'до', 'ед.изм.', 'шаг', 'тип'
            parameter_tile_length = len(mini_table_header)
            column = start_column
            for parameter in parameters:
                sheet.cell(row=row, column=column).value = parameter  # название параметра
                set_cell_style(sheet, row, column, 'parameter_title')

                for i, item in enumerate(mini_table_header):
                    sheet.cell(row=row + 1, column=column + i).value = item
                    set_cell_style(sheet, row + 1, column + i, 'parameter_table')

                column_delta = parameter_tile_length - 1
                sheet.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column + column_delta)
                column += column_delta + 1


def _table_line_output(table_info: numpy.record, sheet: worksheet, row: int) -> int:
    if table_info:
        sheet.cell(row=row, column=column_index_from_string('C')).value = table_info['C']

        for column in range(1, len(headers["table"]) + 1):
            sheet.cell(row=row, column=column).font = items_styles['table']['font']
            sheet.cell(row=row, column=column).fill = items_styles['table']['fill']
            sheet.cell(row=row, column=column).border = items_styles['table']['border_up']

            sheet.cell(row=row + 1, column=column).font = items_styles['table']['font']
            sheet.cell(row=row + 1, column=column).fill = items_styles['table']['fill']
            sheet.cell(row=row + 1, column=column).border = items_styles['table']['border_down']

        sheet.cell(row=row, column=column_index_from_string('C')).font = basic_colors['grey']

        column_h = column_index_from_string('H')
        end_attributes_column = _attributes_output(table_info['G'], start_column=column_h, sheet=sheet, row=row)

        _parameters_output(table_info['H'], end_attributes_column, sheet, row)

        sheet.row_dimensions[row].height = 12
        sheet.row_dimensions[row + 1].height = 12
        return row + 2
    return row


def count_resources_for_table(table_code: str, df: DataFrame) -> int:
    x = df['I'].value_counts().get(table_code, 0)
    return x




def write_tables(input_file_name: str, output_sheet: worksheet, start_line: int, tables_limit: int = 0):
    """
    Записывает информацию о таблицах.
    :param input_file_name: Имя фала с данными.
    :param output_sheet: Лист, на который выводить данные.
    :param start_line: Строка с которой надо начинить запись.
    :param tables_limit: Сколько таблиц выводить.
    """

    data = SourceData(input_file_name)
    data.info()

    if 0 < tables_limit < len(data.tables.index):
        data.tables = data.tables.loc[:tables_limit]
    row = start_line
    for table in data.tables.to_records(index=False):
        start = time.monotonic_ns()
        amount = count_resources_for_table(table['C'], data.resources)
        if amount > 0:
            row = _table_line_output(table, output_sheet, row=row)
            row = write_resources_for_table(data, output_sheet, row, table)
        print(f"таблица: {table['C']} ресурсов: {amount} time: {((time.monotonic_ns()-start)/1e9)} sec")