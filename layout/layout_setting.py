# https://htmlcolorcodes.com/
# https://htmlcolorcodes.com/color-chart/

from openpyxl.styles import Color, Font, PatternFill, Border, Side

side_none = Side(border_style=None)

headers = {
    'main': ['номер', 'место', 'шифр', 'описание', 'измерение', 'статистика', 'флаг', 'атрибуты/параметры', '', ''],
    'option': ['от', 'до', 'ед.изм.', 'шаг', 'тип',],
    'attribute': ['атрибуты',],
    'table': ['A', 'B', 'C', 'D', 'E', 'F', 'G']
}


width_columns = {'A': 6, 'B': 6, 'C': 10, 'D': 100, 'E': 10, 'F': 10, 'G': 10, 'H': 10, 'I': 10, 'J': 10, 'K': 10, 'L': 10, 'M': 10}


items_styles = {
                'resource': {
                    'font': Font(name='Calibri', color=Color(rgb='00000000'), size=8, bold=False),
                    'fill': PatternFill(patternType="solid", fgColor=Color(rgb='00FFFFFF')),
                    'border': Border(top=Side(border_style=None), bottom=Side(border_style='thin', color=Color(rgb='00E5E7E9')))
                },

                'line': {
                    'font': Font(name='Calibri', color=Color(rgb='00D35400'), size=8, bold=False),
                    'fill': PatternFill(patternType="solid", fgColor=Color(rgb='000DF04E')),
                    'border': Border(top=Side(border_style='thin', color=Color(rgb='00A0A0A0')), bottom=Side(border_style='thin', color=Color(rgb='00A0A0A0')))
                },
                'table': {
                    'font': Font(name='Calibri', color=Color(rgb='0034495E'), size=8, bold=False),
                    'fill': PatternFill(patternType="solid", fgColor=Color(rgb='00FCFBFA')),
                    'border_up': Border(top=Side(border_style='thin', color=Color(rgb='00A0A0A0'))),
                    'border_down': Border(bottom=Side(border_style='thin', color=Color(rgb='00A0A0A0')))
                },
                'main_header': {
                    'font': Font(name='Calibri', color=Color(rgb='0034495E'), size=8, bold=True),
                    'fill': PatternFill(patternType="solid", fgColor=Color(rgb='00F3F5EE')),
                    'border': Border(top=Side(border_style='thin', color=Color(rgb='00A0A0A0')), bottom=Side(border_style='thin', color=Color(rgb='00A0A0A0')))
                },
                'attributes_header': {
                    'font': Font(name='Calibri', color=Color(rgb='0034495E'), size=8, bold=False),
                    'fill': PatternFill(patternType="solid", fgColor=Color(rgb='00F7F7F8')),
                    'border': Border(
                                    top=Side(border_style='thin', color=Color(rgb='00A0A0A0')),
                                    bottom=Side(border_style='thin', color=Color(rgb='00A0A0A0')),
                                    left=Side(border_style='thin', color=Color(rgb='00A0A0A0')),
                                    right=Side(border_style='thin', color=Color(rgb='00A0A0A0'))
                                    )
                },
                'attributes_volume': {
                    'font': Font(name='Calibri', color=Color(rgb='0034495E'), size=8, bold=True),
                    'fill': PatternFill(patternType="solid", fgColor=Color(rgb='00F7F7F8')),
                    'border': Border(
                                    top=Side(border_style='thin', color=Color(rgb='00A0A0A0')),
                                    bottom=Side(border_style='thin', color=Color(rgb='00A0A0A0')),
                                    left=Side(border_style='thin', color=Color(rgb='00A0A0A0')),
                                    right=Side(border_style='thin', color=Color(rgb='00A0A0A0'))
                                    )
                },

                'parameter_title': {
                    'font': Font(name='Calibri', color=Color(rgb='0034495E'), size=8, bold=True),
                    'fill': PatternFill(patternType="solid", fgColor=Color(rgb='00EBF1DE')),    # 00DFFF00
                    'border': Border(
                                    top=Side(border_style='thin', color=Color(rgb='00A0A0A0')),
                                    bottom=Side(border_style='thin', color=Color(rgb='00A0A0A0')),
                                    left=Side(border_style='thin', color=Color(rgb='00A0A0A0')),
                                    right=Side(border_style='thin', color=Color(rgb='00A0A0A0'))
                                    )
                },
                'parameter_table': {
                    'font': Font(name='Calibri', color=Color(rgb='0034495E'), size=8, bold=False),
                    'fill': PatternFill(patternType="solid", fgColor=Color(rgb='00EBF1DE')), # 006495ED
                    'border': Border(
                                    top=Side(border_style='thin', color=Color(rgb='00A0A0A0')),
                                    bottom=Side(border_style='thin', color=Color(rgb='00A0A0A0')),
                                    left=Side(border_style='thin', color=Color(rgb='00A0A0A0')),
                                    right=Side(border_style='thin', color=Color(rgb='00A0A0A0'))
                                    )
                },


}


# own_border = Border(left=Side(border_style=None, color='FF000000'),
#                     right=Side(border_style=None, color='FF000000'),
#                     top=Side(border_style=None, color='FF000000'),
#                     bottom=Side(border_style=None, color='FF000000'),
#                     diagonal=Side(border_style=None, color='FF000000'),
#                     diagonal_direction=0,
#                     outline=Side(border_style=None, color='FF000000'),
#                     vertical=Side(border_style=None, color='FF000000'),
#                     horizontal=Side(border_style=None, color='FF000000')
#                     )






