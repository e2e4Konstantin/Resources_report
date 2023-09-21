from openpyxl.worksheet import worksheet

from layout.layout_setting import items_styles


def set_cell_style(sheet: worksheet, row: int, column: int, style_name: str):
    """ Устанавливает оформление для ячейки """
    sheet.cell(row=row, column=column).font = items_styles[style_name]['font']
    sheet.cell(row=row, column=column).fill = items_styles[style_name]['fill']
    sheet.cell(row=row, column=column).border = items_styles[style_name]['border']
